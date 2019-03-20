#!/usr/bin/env python3
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from tkinter import ttk
from openpyxl import Workbook
from openpyxl import load_workbook
import configparser
import requests
import xmltodict
import csv
import os
import datetime

# local modules
import alma

# main program ################################################################
def main(*args):
    input_file = gui.openfile()
    if input_file == "":
        return
        
    # clear tree view and progress bar
    gui.clear_tree()
    gui.progress_bar["value"] = 0
        
    # check file extension
    file_extension = os.path.splitext(input_file)[1]
    
    # if plain text file, convert to XLSX
    if file_extension.upper() == ".TXT" or file_extension.upper() == ".CSV":
        # set delimiters
        if file_extension.upper() == ".TXT":
            delimiter = "\t"
        if file_extension.upper() == ".CSV":
            delimiter = ","
        
        # convert to XLSX
        with open(input_file, newline='\n', encoding='utf-8') as plain_text_file:
            reader = csv.reader(plain_text_file, delimiter=delimiter, quotechar='"')
            
            wb = Workbook()
            ws = wb.active
            
            for row in reader:
                ws.append(row)
        
        wb.save("Converted_PlainText_File.xlsx")
        input_file_converted = "Converted_PlainText_File.xlsx"
        
    # if already XLSX file
    if file_extension.upper() == ".XLSX":
        input_file_converted = input_file
       
    # open excel file
    wb = load_workbook(filename=input_file_converted)
    ws = wb.get_sheet_by_name(wb.sheetnames[0])
    row_count = ws.max_row - 1
    
    # find barcode column
    barcode_column = find_barcode_column(ws)
    if barcode_column == None:
        gui.msgbox("No Barcode column found.")
        return
    
    # loop through barcodes
    counter = 0
    for cell in ws[barcode_column]:
        # skip header
        if config.barcode_column_header != "" and counter == 0:
            counter += 1
            gui.insert_text(counter, ("n/a", "n/a", "Skipped", "Spreadsheet header."), 'attention')
            continue
        
        # get item record
        barcode = str(cell.value)
        
        # clean up prefixes and suffixes
        if config.prefix_trim > 0:
            barcode = barcode[config.prefix_trim:]
        if config.suffix_trim > 0:
           barcode = barcode[:-config.suffix_trim]
        
        # get information from item record in Alma
        item = alma.item_record(barcode, config.apikey)
        if item.found == False:
            counter += 1
            gui.insert_text(counter, (barcode, "n/a", "Error", item.error_msg), 
                                      'error')
            continue
        
        # return item
        scan_return = alma.ret()
        scan_return.post(config.apikey, config.library, config.circ_desk, 
                         config.register_in_house_use, item.mms_id, 
                         item.holding_id, item.pid, item.xml)
        if scan_return.successful == False:
            counter += 1
            gui.insert_text(counter, (barcode, item.title, "Error", 
                         scan_return.error_msg), 'error')
            continue
        
        if scan_return.successful == True:
            if "Queue: 0" in scan_return.additional_info:
                counter += 1
                
                # if file only has a single line of barcodes, you will get a div by zero error
                try:
                    increment = 100 / row_count
                except:
                    increment = 100 / 1
                
                gui.insert_text(counter, (barcode, item.title, "Returned", 
                                          scan_return.additional_info), 'success')
                gui.progress_bar.step(increment)
            
            if "Queue: 0" not in scan_return.additional_info:
                counter += 1
                increment = 100 / row_count
                gui.insert_text(counter, (barcode, item.title, "Returned-Queue", 
                                          scan_return.additional_info), 'attention')
                gui.progress_bar.step(increment)
        
    # delete source file
    if config.delete_barcode_file.upper() == "TRUE":
        try:
            os.remove(input_file)
        except OSError as e:
            gui.msgbox(e)
    
    # finish
    gui.progress_bar["value"] = 100
    gui.msgbox("Done.")
    
# functions ###################################################################
def find_barcode_column(ws):
    barcode_column = None
        
    # if no header specified
    if config.barcode_column_header == "":
        barcode_column = "A"
    
    # if header is specified
    if config.barcode_column_header != "":
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.value.upper() == config.barcode_column_header.upper():
                        barcode_column = cell.column
                        break
                        
    # finish
    return barcode_column
    
# configs #####################################################################
class configs:
    def __init__(self, configfile):
        self.configs = configs

        c_dict = configparser.ConfigParser()
        c_dict.read(configfile)
        
        self.version                 = c_dict['misc']['version']
        self.delete_barcode_file     = c_dict['misc']['delete_barcode_file']

        self.key                     = c_dict['apikey']['key']
        
        self.download_directory      = c_dict['spreadsheet']['spreadsheet_directory'].replace('\\', '//')
        self.barcode_column_header   = c_dict['spreadsheet']['barcode_column_header']
        
        self.library                 = c_dict['alma']['library']
        self.circ_desk               = c_dict['alma']['circ_desk']
        self.register_in_house_use   = c_dict['alma']['register_in_house_use'] 
        
        self.prefix_trim             = int(c_dict['barcodes']['prefix_trim'])
        self.suffix_trim             = int(c_dict['barcodes']['suffix_trim'])
        
        self.log_directory           = c_dict['log']['log_directory'].replace('\\', '//')
        
        # set apikey
        # -will first search for a matching environmental variable key. 
        # -if not found will set apikey as written in config.ini entry.
        try:
            self.apikey = os.environ[self.key]
        except KeyError:
            self.apikey = self.key
        
# gui #########################################################################
class gui:
    def __init__(self, master):
        self.master = master
        
        master.title("RoboReturns "+config.version)
        master.resizable(0, 0)
        master.minsize(width=1150, height=500)
        master.maxsize(width=1150, height=900)
        master.iconbitmap(".\images\logo_small.ico")

        logo = PhotoImage(file=".\images\logo_large.png")
        self.logo = Label(image=logo)
        self.logo.image = logo
        self.logo.pack()
        
        # frames
        self.top_frame = Frame(master)
        self.top_frame.pack(side='top', fill='both', expand=False)
        
        self.run_button = Button(self.top_frame, text="OPEN FILE AND RUN", font="Arial 14", command=main, relief="groove")
        self.run_button.pack(fill='both', side='left', expand=True)
        
        self.save_img = PhotoImage(format = 'png', file= '.\images\save_icon.png')
        self.save_button = Button(self.top_frame, text="SAVE LOG", image=self.save_img, font="Arial 14", command=self.save_log, relief="groove")
        self.save_button.pack(fill='both', side='right', expand=False)
        
        self.mid_frame = Frame(master)
        self.mid_frame.pack(side='top', fill='both', expand=True)
        
        # tree view
        self.tree = ttk.Treeview(self.mid_frame, height=15)
        style = ttk.Style()
        style.theme_use('clam')
        
        # headings
        self.tree['columns'] = ('barcode', 'title', 'status', 'info')
        self.tree.heading('#0', text='#', anchor='w')
        self.tree.heading('barcode', text='Barcode', anchor="w")
        self.tree.heading('title', text='Title', anchor="w")
        self.tree.heading('status', text='Status', anchor="w")
        self.tree.heading('info', text='Additional Info', anchor="w")
        
        self.tree.column("#0", width=40)
        self.tree.column("barcode", width=100)
        self.tree.column("title", width=300)
        self.tree.column("status", width=100)
        self.tree.column("info", width=900)
        
        self.tree.pack(fill="both", expand=False, side="left")
        
        # scrollbar
        v_scrollbar = ttk.Scrollbar(self.mid_frame, orient="vertical", 
                                    command=self.tree.yview)
        v_scrollbar.place(x=1136, y=26, height=303)
        self.tree.configure(yscrollcommand=v_scrollbar.set)
       
        # tags
        self.tree.tag_configure('error', background='pink')
        self.tree.tag_configure('success', background='white')
        self.tree.tag_configure('attention', background='khaki')
       
        # progressbar
        style.configure("red.Horizontal.TProgressbar", foreground='red', 
                        background='#2381df')
        self.progress_bar = ttk.Progressbar(master, style="red.Horizontal.TProgressbar", 
                                            orient='horizontal', mode='determinate')
        self.progress_bar.pack(fill="both", expand=False, side="top")
        
    def msgbox(self, msg):
        messagebox.showinfo("Attention", msg)
    
    def openfile(self):
        self.filename =  filedialog.askopenfilename(initialdir = config.download_directory, 
                           title = "Select file", 
                           filetypes = (("EXCEL, TXT, or CSV","*.xlsx *.txt *.csv"),("all files","*.*")))
        filename = self.filename
        return filename
        
    def clear_tree(self):
        self.tree.delete(*self.tree.get_children())
        
    def insert_text(self, counter, msg, tags):
        self.tree.insert("", "end", text=counter, values=(msg), tags=tags)
        self.tree.yview_moveto(1)
        root.update()
        
    def save_log(self):
        current_date = datetime.datetime.now()
        current_date_formatted = current_date.strftime("%Y-%m-%d__%H%M%S")
        saved_log = open(f"{config.log_directory}robo_returns_log_{current_date_formatted}.csv", "w", encoding="utf-8", newline='')
        children = self.tree.get_children()
        for child in children:
            list = self.tree.item(child)["values"]
            w = csv.writer(saved_log, quoting=csv.QUOTE_ALL)
            w.writerow(list)
        saved_log.close()
        self.msgbox("LOG SAVED SUCCESFULLY.")

# toplevel #########################################################################
config = configs('config.ini')

# gui
root = Tk()
gui = gui(root)
root.mainloop()